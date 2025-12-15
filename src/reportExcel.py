from pathlib import Path
import pandas as pd

from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference


def export_excel(
    total: float,
    por_mes: pd.DataFrame,
    por_categoria: pd.DataFrame,
    por_medio_pago: pd.DataFrame,
    top_gastos: pd.DataFrame,
    out_path: Path,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # 1) Crear Excel con pandas (hojas)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame({"total": [total]}).to_excel(writer, index=False, sheet_name="resumen")
        por_mes.to_excel(writer, index=False, sheet_name="por_mes")
        por_categoria.to_excel(writer, index=False, sheet_name="por_categoria")
        por_medio_pago.to_excel(writer, index=False, sheet_name="por_medio_pago")
        top_gastos.to_excel(writer, index=False, sheet_name="top_10_gastos")

    # 2) Reabrir con openpyxl para agregar gráficos
    wb = load_workbook(out_path)

    # ---- Gráfico de torta: gastos por categoría (hoja por_categoria) ----
    ws_cat = wb["por_categoria"]

    grafico_torta = PieChart()
    grafico_torta.title = "Gastos por categoría"

    data_cat = Reference(
        ws_cat,
        min_col=2,      # columna "monto"
        min_row=2,      # desde primera fila de datos
        max_row=ws_cat.max_row,
    )
    labels_cat = Reference(
        ws_cat,
        min_col=1,      # columna "categoria"
        min_row=2,
        max_row=ws_cat.max_row,
    )

    grafico_torta.add_data(data_cat, titles_from_data=False)
    grafico_torta.set_categories(labels_cat)

    ws_cat.add_chart(grafico_torta, "E2")

    # ---- Gráfico de barras: gastos por mes (hoja por_mes) ----
    ws_mes = wb["por_mes"]

    grafico_barras = BarChart()
    grafico_barras.title = "Gastos por mes"
    grafico_barras.y_axis.title = "Monto"
    grafico_barras.x_axis.title = "Mes"

    data_mes = Reference(
        ws_mes,
        min_col=2,      # columna "monto"
        min_row=1,      # incluye encabezado para título
        max_row=ws_mes.max_row,
    )
    categorias_mes = Reference(
        ws_mes,
        min_col=1,      # columna "mes"
        min_row=2,
        max_row=ws_mes.max_row,
    )

    grafico_barras.add_data(data_mes, titles_from_data=True)
    grafico_barras.set_categories(categorias_mes)

    ws_mes.add_chart(grafico_barras, "E2")

    wb.save(out_path)
